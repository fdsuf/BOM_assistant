import tkinter as tk
from tkinter import filedialog
import os,re,time,threading
from openpyxl import load_workbook
from datetime import datetime

# 获取脚本所在目录
script_dir = os.path.dirname(os.path.abspath(__file__))

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    file1_entry.delete(0, tk.END)
    file1_entry.insert(0, file_path)

def select_txt_file():
    file_path = filedialog.askopenfilename(filetypes=[("Text files", ".txt")])
    file2_entry.delete(0, tk.END)
    file2_entry.insert(0, file_path)

log_text = None
process_i = None

def initialize_log_text(root):
    global log_text
    if log_text is None:
        # 创建滚动文本框显示日志，只需创建一次
        log_text = tk.Text(root, width=60, height=11, wrap="word")
        log_text.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(root, command=log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        log_text.config(yscrollcommand=scrollbar.set)
        
        log_text.insert(tk.END, "操作步骤：\n")
        log_text.insert(tk.END, "  1.原理图导出为txt格式文件\n")
        log_text.insert(tk.END, "  2.选择BOM文件\n")
        log_text.insert(tk.END, "  3.选择第1步的txt文件\n")
        log_text.insert(tk.END, "  4.点击‘运行’，结束后再把txt文件导入原理图中\n")
        log_text.insert(tk.END, "\n注意：\n  导入时要先打开一个空白的原理图\n")

def BOM_process_thread(excel_file_path_widget):
    global process_i
    print(f"\n加载BOM文件... \n")
    # 加载Excel工作簿
    wb = load_workbook(filename=excel_file_path_widget, keep_vba=False)
    sheet = wb[wb.sheetnames[0]]#.active
    print(f"\n加载完成... \n")
    # 查找列标题
    start_row = None
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if "参考标记(贴片位置)" in row and "物料名称" in row:
            start_row = [row.index("参考标记(贴片位置)"), row.index("物料名称")]
            break

    if not start_row:
        print("未找到列标题")
        exit()


    # 读取数据并处理
    data = []
    # 从找到列标题的下一行开始读取
    for row in sheet.iter_rows(min_row=sheet.max_row+1 if start_row is None else start_row[0] + 1, values_only=True, max_row=1000,max_col=20):
        try:
            prefixes = ('C','R','L')
            # 根据索引获取数据
            if row[start_row[0]] is not None and row[start_row[0]].startswith(prefixes):
                ref_mark = row[start_row[0]]
                material_name = row[start_row[1]]
            else:
                ref_mark = row[start_row[0]]
                material_name = row[start_row[1]+5]
            # 拼接数据并替换逗号，删除特定字样Ω
            if material_name == 'steel_frame' or material_name == 'BRACKET' or ref_mark == None:
                #print(material_name,ref_mark)
                continue
            processed_data = f"{ref_mark} value {material_name}".replace(",", " ").replace("CAP", "").replace("IND", "").replace("RES", "").replace("Power", "").replace("Ω", "R")
            
            data.append(processed_data)
        except IndexError:
            # 如果索引超出范围，跳过当前行
            continue

    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 构建txt文件的完整路径
    output_file_path = os.path.join(script_dir, "processed_data.txt")
    print('output_file_path:',output_file_path)
    # 写入到txt文件
    with open(output_file_path, 'w', encoding='utf-8') as file:
        for item in data:
            file.write("%s\n" % item.strip())

    process_i = 1
    print('把位号和值写入到sch_data中...\n')

      
def sch_process_thread(txt_file_path_widget):
    global process_i
    while True:
        time.sleep(0.5)
        print('Wait:',process_i,datetime.now())
        if process_i == 1:
            break
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_file_path = os.path.join(script_dir, "processed_data.txt")
    components_values = {}
    with open(output_file_path, 'r', encoding='utf-8') as file2:
        for line in file2:
            # 使用正则表达式查找可能存在的多个标识符和随后的value值
            match = re.search(r'((?:\w+\s+)*)(value\s.+)', line)
            
            if match:
                # 抽取所有标识符（它们之间由空格分隔）
                identifiers = match.group(1).split()
                
                # 抽取value值
                value_match = re.search(r'value(\s.+)', match.group(2))
                
                if value_match:
                    value = value_match.group(1).strip()
                    # 将每个标识符及其对应的value值存储到字典中
                    for identifier in identifiers:
                        components_values[identifier.strip()] = value
                        

    # 打印出解析得到的字典内容，以验证是否正确
    #for comp, val in components_values.items():
     #   print(f"Component {comp} has value {val}")


    with open(txt_file_path_widget, 'r', encoding='gbk', errors='ignore') as file1:
        file1_lines = file1.readlines()


    # 用于存储更新后的内容
    updated_lines = []

    # 先将第一个文档的所有Value值替换为NC
    for line in file1_lines:
        # 替换Value为NC，这里我们使用非贪婪匹配+来匹配Value后面的空格
        updated_line = re.sub(r'"Value".*', r'"Value" NC', line)
        
        updated_lines.append(updated_line)
    with open(txt_file_path_widget, 'w', encoding='gbk', errors='ignore') as file1:
        file1.writelines(updated_lines)

    # 读取第一个文件并逐行分析
    with open(txt_file_path_widget, 'r', encoding='gbk', errors='ignore') as file1:
        file1_lines = file1.readlines()


    # 用于存储更新后的内容
    updated_lines = []

    component_found = False

    for line in file1_lines:
        if component_found:
            if line == '':
                component_found = False

            # 如果当前行包含"Value"，并且我们已经找到了对应的组件
            if '"Value"' in line:
                # 替换Value后面的值
                #print(line)
                original_value = line.split('"Value"')[-1].strip().split()[-1]
                new_value = components_values.get(component_get)  # 假设我们要找的组件是L3124
                line = line.replace(original_value, new_value)
                component_found = False  # 重置标志
        if any(component[0]+' ' in line for component in components_values.items()):
            
            for component in components_values.items():
                if (component[0]+' ') in line:
                    component_get = component[0]
                    #print('Found:',component_get)

            component_found = True  # 找到组件，下一行将替换Value
            found_i = 0

        # 添加当前行到更新后的列表中
        updated_lines.append(line)

    # 将更新后的内容写回到第一个文件
    with open(txt_file_path_widget, 'w') as file1:
        file1.writelines(updated_lines)
    
    process_i = 2
    print("写入完成...")      
      
def log_process_thread():
    global process_i
    log_i = 0
    while True:
        if process_i == None and log_i == 0:
            #threading.Thread(target=update_log, args=(log_text, '\n读取BOM文件...个别表格可以要读取较慢，不要重复点运行.\n'), daemon=True).start()
            update_log(log_text, '\n读取BOM文件...个别表格可能读取较慢，不要重复点运行.\n')
            log_i = 1
        elif (process_i == None and log_i == 1) or (process_i == 1 and log_i == 2):
            time.sleep(1)
            update_log(log_text, '*')
        elif process_i == 1 and log_i == 1:
            #threading.Thread(target=update_log, args=(log_text, '\n把位号和值写入到sch_data中...\n'), daemon=True).start()
            update_log(log_text, '\n把位号和值写入到sch_data中...\n')
            log_i = 2
        elif process_i == 2 and log_i == 2:
            #threading.Thread(target=update_log, args=(log_text, '\n把位号和值写入到sch_data中...\n'), daemon=True).start()
            update_log(log_text, '\n写入完成...\n')
            log_i = 0
            break
            
      
def run_scripts2():
    global log_text
    global process_i
    process_i = None
    if log_text is None:
        initialize_log_text(root)
    
    # 清除之前的日志
    log_text.delete(1.0, tk.END)
    
    
    excel_file_path = file1_entry.get()
    txt_file_path = file2_entry.get()
    # 创建滚动文本框显示日志
    #log_text = st.ScrolledText(root, width=40, height=10)
    #log_text.pack()
    if excel_file_path and txt_file_path:
        # 显示运行中状态
        #status_label.config(text="运行中...")
        
        # 运行 BOM_process.py 脚本
        if process_i == None:
            
            threading.Thread(target=log_process_thread, args=(), daemon=True).start()
            threading.Thread(target=BOM_process_thread, args=(excel_file_path,), daemon=True).start()
            
            
            threading.Thread(target=sch_process_thread, args=(txt_file_path,), daemon=True).start()


def update_log(log_text_widget, output_fd):
    line = output_fd
    # 安排在主线程稍后更新UI
    log_text_widget.after_idle(lambda: log_text_widget.insert(tk.END, line))
    log_text_widget.after_idle(log_text_widget.see, tk.END)  # 自动滚动到日志的最后


# 创建主窗口
root = tk.Tk()
root.geometry("400x300")  # 设置窗口大小
root.title("BOM助手 - By Coosea")
#root.iconbitmap("F:\\射频匹配导入\\exe3\\BOM3.ico")
# 调用 initialize_log_text 来初始化日志窗口


# 创建一个 Frame 容器，使用 pack 布局管理器
top_frame = tk.Frame(root)
top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
# 创建一个 Frame 容器，使用 pack 布局管理器
top_frame2 = tk.Frame(root)
top_frame2.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
# 创建一个 Frame 容器，使用 pack 布局管理器
top_frame3 = tk.Frame(root)
top_frame3.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)


# 创建第一个文件选择框及按钮
file1_label = tk.Label(top_frame, text="选择 BOM 文件:")
file1_label.pack(side=tk.LEFT,anchor=tk.W)
file1_entry = tk.Entry(top_frame)
file1_entry.pack(side=tk.LEFT,anchor=tk.W, padx=10, pady=10)
select_excel_button = tk.Button(top_frame, text="选择文件", command=select_excel_file)
select_excel_button.pack(side=tk.LEFT)

# 创建第二个文件选择框及按钮
file2_label = tk.Label(top_frame2, text="选择 TXT 文件:")
file2_label.pack(side=tk.LEFT,anchor=tk.W)
file2_entry = tk.Entry(top_frame2)
file2_entry.pack(side=tk.LEFT,anchor=tk.W, padx=(18,10), pady=10)
select_txt_button = tk.Button(top_frame2, text="选择文件", command=select_txt_file)
select_txt_button.pack(side=tk.LEFT)


# 创建运行按钮和退出按钮
quit_button = tk.Button(top_frame3, text="退 出", command=root.destroy)
quit_button.pack(side=tk.RIGHT, padx=5, pady=5)
quit_button['bg'] = '#E8E8E8'
run_button = tk.Button(top_frame3, text="运 行", command=run_scripts2)
run_button.pack(side=tk.RIGHT, padx=5, pady=5)
run_button['bg'] = '#DDA0DD'

# 创建用于显示状态的标签
#status_label = tk.Label(root, text="就绪")
#status_label.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
initialize_log_text(root)
# 主循环
root.mainloop()